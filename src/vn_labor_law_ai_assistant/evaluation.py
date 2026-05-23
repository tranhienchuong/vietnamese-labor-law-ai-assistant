from __future__ import annotations

from dataclasses import asdict, dataclass, field
import csv
import json
from pathlib import Path
import re
from typing import Iterable, Sequence

from .corpus_pipeline import normalize_for_matching
from .core.config import load_settings
from .evaluation_citation_matcher import (
    CitationRef,
    citation_contains,
    citation_matches,
    normalize_match_mode,
    normalize_vietnamese_citation,
    parse_citation,
)


WORKBOOK_SHEET_NAME = "golden_benchmark"
WORKBOOK_RESULTS_SHEET_NAME = "evaluation_results"
BENCHMARK_JSONL_NAME = "golden_benchmark_100_answered_v1.jsonl"
RESULTS_COLUMNS = (
    "id",
    "model_version",
    "retrieval_hit_at_5",
    "retrieval_first_relevant_rank",
    "retrieval_reciprocal_rank",
    "citation_correct",
    "citation_document_correct",
    "citation_provision_correct",
    "citation_article_correct",
    "citation_supports_answer",
    "answer_correct",
    "legal_issue_classification_correct",
    "legal_reasoning_score_1_5",
    "missing_information_score_0_2",
    "hallucination_flag",
    "hallucination_types",
    "abstention_correct",
    "groundedness_score_1_5",
    "clarity_score_1_5",
    "format_score_1_5",
    "final_score_10",
    "evaluator",
    "comments",
    "skill_tag",
    "question",
    "expected_citations",
    "expected_citations_in_scope",
    "expected_citations_out_of_scope",
    "case_scope",
    "retrieved_citations",
    "generated_answer",
    "generated_legal_basis",
    "generated_evidence_quotes",
    "insufficient_context",
)

JUDGE_JSON_SCHEMA = {
    "type": "object",
    "properties": {
        "answer_correct": {
            "type": "string",
            "enum": ["yes", "partial", "no"],
        },
        "legal_issue_classification_correct": {
            "type": "string",
            "enum": ["yes", "partial", "no"],
        },
        "legal_reasoning_score_1_5": {"type": "integer", "minimum": 1, "maximum": 5},
        "missing_information_score_0_2": {"type": "integer", "minimum": 0, "maximum": 2},
        "citation_supports_answer": {
            "type": "string",
            "enum": ["yes", "partial", "no"],
        },
        "groundedness_score_1_5": {"type": "integer", "minimum": 1, "maximum": 5},
        "clarity_score_1_5": {"type": "integer", "minimum": 1, "maximum": 5},
        "format_score_1_5": {"type": "integer", "minimum": 1, "maximum": 5},
        "hallucination_types": {
            "type": "array",
            "items": {
                "type": "string",
                "enum": ["legal_basis", "rule", "fact"],
            },
        },
        "comments": {"type": "string"},
    },
    "required": [
        "answer_correct",
        "legal_issue_classification_correct",
        "legal_reasoning_score_1_5",
        "missing_information_score_0_2",
        "citation_supports_answer",
        "groundedness_score_1_5",
        "clarity_score_1_5",
        "format_score_1_5",
        "hallucination_types",
        "comments",
    ],
    "additionalProperties": False,
}

JUDGE_SYSTEM_PROMPT = """Bạn là giám khảo benchmark cho trợ lý pháp lý lao động.

Hãy so sánh GENERATED_ANSWER với GOLD_ANSWER và chấm thật nghiêm khắc.

Tiêu chí:
1. answer_correct:
- yes: trả lời đúng ý chính và không có sai sót pháp lý đáng kể.
- partial: đúng một phần nhưng thiếu ý quan trọng, quá chung chung, hoặc có lỗi nhẹ.
- no: sai trọng yếu, mâu thuẫn với đáp án chuẩn, hoặc trả lời lạc đề.
2. legal_issue_classification_correct:
- yes: phân loại đúng chế định pháp lý trọng tâm, ví dụ đơn phương chấm dứt, sa thải kỷ luật, thỏa thuận chấm dứt, trợ cấp thôi việc/mất việc.
- partial: nhận ra một phần chế định nhưng còn lẫn hoặc thiếu nhánh pháp lý quan trọng.
- no: nhầm chế định pháp lý trọng yếu, ví dụ nhầm sa thải với đơn phương chấm dứt hoặc nhầm trợ cấp thôi việc với trợ cấp mất việc làm.
3. legal_reasoning_score_1_5:
- 5: phân loại đúng chế định, nêu đủ điều kiện, ngoại lệ, thủ tục bắt buộc và hậu quả pháp lý chính.
- 4: đúng chính, chỉ thiếu chi tiết nhỏ.
- 3: kết luận đúng nhưng thiếu điều kiện quan trọng.
- 2: quá chung chung hoặc nhầm một phần chế định.
- 1: sai chế định pháp lý.
4. missing_information_score_0_2:
- 2: nhận diện đúng dữ kiện thiếu và kết luận có điều kiện; nếu câu hỏi đã đủ dữ kiện thì không tự thêm điều kiện không cần thiết.
- 1: có nói thiếu thông tin nhưng chưa rõ hoặc chưa gắn với kết luận.
- 0: kết luận chắc chắn khi thiếu dữ kiện quan trọng, hoặc từ chối/đòi thêm dữ kiện không hợp lý.
5. citation_supports_answer:
- yes: citation được nêu thật sự hỗ trợ kết luận.
- partial: citation có liên quan nhưng chưa đủ cho toàn bộ kết luận.
- no: citation không hỗ trợ kết luận, nêu cho có, hoặc sai căn cứ.
6. groundedness_score_1_5:
- 5: mọi khẳng định pháp lý đều bám sát EXPECTED_CITATIONS_IN_SCOPE và không vượt quá RETRIEVED_CITATIONS.
- 3: có ý đúng nhưng vẫn còn diễn giải rộng hơn evidence được cấp.
- 1: có khẳng định pháp lý, ngoại lệ, hoặc citation không được hỗ trợ bởi evidence được cấp.
7. clarity_score_1_5: độ mạch lạc, dễ hiểu, ngắn gọn đúng mức.
8. format_score_1_5: đúng format trả lời ngắn gọn, có cấu trúc hợp lý, không lan man.
9. hallucination_types:
- legal_basis: bịa hoặc dùng sai căn cứ/citation.
- rule: nêu quy tắc pháp lý không có trong evidence hoặc trái đáp án chuẩn.
- fact: tự thêm dữ kiện không có trong câu hỏi/context.
- trả [] nếu không có hallucination rõ.
10. comments: tối đa 2 câu, nêu rõ điểm mạnh và điểm yếu chính.

Lưu ý:
- Chỉ được chấm dựa trên thông tin được cung cấp trong prompt này.
- Nếu benchmark yêu cầu abstain và GENERATED_ANSWER từ chối hợp lý, không nên chấm thấp chỉ vì thiếu kết luận.
- Thiếu dữ kiện không đồng nghĩa phải từ chối hoàn toàn: câu trả lời tốt nên nêu nguyên tắc, điều kiện còn thiếu và các nhánh kết luận nếu phù hợp.
- Nếu GENERATED_ANSWER đưa ra thông tin pháp lý nằm ngoài EXPECTED_CITATIONS_IN_SCOPE hoặc trái với RETRIEVED_CITATIONS, phải coi đó là groundedness thấp.
- Không tự chấm final_score_10; hệ thống sẽ tính điểm tổng hợp bằng công thức riêng.
- Trả đúng JSON theo schema."""

LEGAL_ARTICLE_RE = re.compile(r"\bdieu\s+(?P<value>\d+[a-z]?)")
LEGAL_CLAUSE_RE = re.compile(r"\bkhoan\s+(?P<value>\d+)")
LEGAL_POINT_GROUP_RE = re.compile(
    r"\bdiem\s+(?P<value>[a-z](?:\s*(?:,|va)\s*[a-z])*)"
)
TOP_LEVEL_CITATION_SPLIT_RE = re.compile(r"\s*;\s*")

DOCUMENT_FAMILY_LABELS = {
    "bo_luat_2019": "Bộ luật Lao động 2019",
    "bo_luat_2012": "Bộ luật Lao động 2012",
    "nghi_dinh_145": "Nghị định 145/2020/NĐ-CP",
    "nghi_dinh_12_2022": "Nghị định 12/2022/NĐ-CP",
    "nghi_dinh_115_2015": "Nghị định 115/2015/NĐ-CP",
    "nghi_dinh_158_2025": "Nghị định 158/2025/NĐ-CP",
    "nghi_dinh_57_2026": "Nghị định 57/2026/NĐ-CP",
    "thong_tu_09_2020": "Thông tư 09/2020/TT-BLĐTBXH",
    "thong_tu_10_2020": "Thông tư 10/2020/TT-BLĐTBXH",
    "luat_bhxh_2014": "Luật Bảo hiểm xã hội 2014",
    "luat_bhxh_2024": "Luật Bảo hiểm xã hội 2024",
    "luat_cong_doan_2024": "Luật Công đoàn 2024",
    "luat_giao_dich_dien_tu_2023": "Lu\u1eadt Giao d\u1ecbch \u0111i\u1ec7n t\u1eed 2023",
    "bo_luat_dan_su_2015": "Bộ luật Dân sự 2015",
    "bo_luat_to_tung_dan_su_2015": "Bộ luật Tố tụng dân sự 2015",
    "luat_kinh_doanh_bao_hiem_2022": "Luật Kinh doanh bảo hiểm 2022",
    "luat_phuc_hoi_pha_san_2025": "Luật Phục hồi, phá sản 2025",
    "cong_van_1198_bhxh": "Công văn 1198/CTL&BHXH-BHXH",
    "bo_tu_phap_guidance": "Giải thích của Bộ Tư pháp",
    "judicial_practice": "Thực tiễn xét xử",
    "thuvienphapluat_article": "THƯ VIỆN PHÁP LUẬT",
}
DOCUMENT_FAMILY_SIGNATURES = {
    "bo_luat_2019": (
        "bo luat lao dong 2019",
        "labor code 2019",
        "45 2019 qh 14",
        "45 2019 qh14",
    ),
    "bo_luat_2012": (
        "bo luat lao dong 2012",
        "labor code 2012",
    ),
    "nghi_dinh_145": (
        "nghi dinh 145 2020 nd cp",
        "145 2020 nd cp",
    ),
    "nghi_dinh_12_2022": (
        "nghi dinh 12 2022 nd cp",
        "12 2022 nd cp",
    ),
    "nghi_dinh_115_2015": (
        "nghi dinh 115 2015 nd cp",
        "115 2015 nd cp",
    ),
    "nghi_dinh_158_2025": (
        "nghi dinh 158 2025 nd cp",
        "158 2025 nd cp",
    ),
    "nghi_dinh_57_2026": (
        "nghi dinh 57 2026 nd cp",
        "57 2026 nd cp",
    ),
    "thong_tu_09_2020": (
        "thong tu 09 2020 tt bldtbxh",
        "09 2020 tt bldtbxh",
    ),
    "thong_tu_10_2020": (
        "thong tu 10 2020 tt bldtbxh",
        "10 2020 tt bldtbxh",
    ),
    "luat_bhxh_2014": (
        "luat bao hiem xa hoi 2014",
        "luat bhxh 2014",
    ),
    "luat_bhxh_2024": (
        "luat bao hiem xa hoi 2024",
        "luat bhxh 2024",
    ),
    "luat_cong_doan_2024": ("luat cong doan 2024",),
    "luat_giao_dich_dien_tu_2023": (
        "luat giao dich dien tu 2023",
        "luat giao dich dien tu",
    ),
    "bo_luat_dan_su_2015": (
        "bo luat dan su 2015",
        "blds 2015",
    ),
    "bo_luat_to_tung_dan_su_2015": (
        "bo luat to tung dan su 2015",
        "blttds 2015",
    ),
    "luat_kinh_doanh_bao_hiem_2022": ("luat kinh doanh bao hiem 2022",),
    "luat_phuc_hoi_pha_san_2025": (
        "luat phuc hoi pha san 2025",
        "luat phuc hoi va pha san 2025",
    ),
    "cong_van_1198_bhxh": ("cong van 1198 ctl bhxh bhxh",),
    "bo_tu_phap_guidance": ("giai thich cua bo tu phap",),
    "judicial_practice": ("thuc tien xet xu",),
    "thuvienphapluat_article": ("thu vien phap luat",),
}


def require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for benchmark workbook import.") from exc
    return load_workbook


@dataclass(frozen=True)
class BenchmarkCase:
    id: str
    category: str
    subtopic: str
    difficulty: str
    question_type: str
    question: str
    scenario: str
    gold_issue: str
    gold_citation_primary: str | None
    gold_citation_secondary: str | None
    gold_answer_short: str
    gold_answer_full: str
    abstain_required: bool
    missing_information: str | None
    source_document: str | None
    source_url: str | None
    annotator: str | None
    review_status: str | None
    notes: str | None
    skill_tag: str | None = None
    gold_citations: tuple[str, ...] = field(default_factory=tuple)

    def __post_init__(self) -> None:
        raw_citations: Iterable[str | None]
        if self.gold_citations:
            raw_citations = tuple(self.gold_citations)
        else:
            raw_citations = (
                self.gold_citation_primary,
                self.gold_citation_secondary,
            )
        object.__setattr__(
            self,
            "gold_citations",
            expand_expected_citations(raw_citations),
        )
        object.__setattr__(
            self,
            "skill_tag",
            self.skill_tag or infer_skill_tag(self),
        )


@dataclass(frozen=True)
class JudgeScore:
    answer_correct: str
    legal_issue_classification_correct: str
    legal_reasoning_score_1_5: int
    missing_information_score_0_2: int
    citation_supports_answer: str
    groundedness_score_1_5: int
    clarity_score_1_5: int
    format_score_1_5: int
    hallucination_types: tuple[str, ...]
    comments: str
    raw_content: str


def coerce_optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def coerce_required_text(value: object, field_name: str) -> str:
    text = coerce_optional_text(value)
    if text is None:
        raise ValueError(f"Missing required benchmark field: {field_name}")
    return text


def parse_yes_no_flag(value: object) -> bool:
    text = normalize_for_matching(str(value or ""))
    return text in {"yes", "y", "true", "1"}


def contains_placeholder_text(value: object) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return (
        "[article]" in text
        or "[clause]" in text
        or "example.com" in text
        or "replace the placeholder" in text
    )


def is_placeholder_benchmark_row(row_values: Sequence[object]) -> bool:
    return any(contains_placeholder_text(value) for value in row_values)


def infer_skill_tag(case: BenchmarkCase) -> str:
    text = normalize_for_matching(
        " ".join(
            (
                case.category,
                case.subtopic,
                case.question_type,
                case.question,
                case.gold_issue,
            )
        )
    )
    if case.missing_information or case.abstain_required:
        return "missing_fact_handling"
    if any(
        token in text
        for token in (
            "sa thai",
            "don phuong",
            "thoa thuan cham dut",
            "unilateral termination",
            "resignation",
            "dismissal",
            "mutual termination",
        )
    ):
        return "legal_classification"
    if any(token in text for token in ("thu tuc", "quy trinh", "ho so", "thoi han", "bao truoc")):
        return "procedure_checking"
    if any(token in text for token in ("tro cap", "boi thuong", "tien luong", "thanh toan", "bao hiem")):
        return "remedy_calculation"
    return "rule_lookup"


def unique_preserve_order(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = value.strip()
        if not cleaned:
            continue
        if cleaned in seen:
            continue
        seen.add(cleaned)
        ordered.append(cleaned)
    return tuple(ordered)


def normalize_signature_text(text: str) -> str:
    normalized = normalize_for_matching(text)
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def citation_document_families(text: str) -> tuple[str, ...]:
    normalized = normalize_signature_text(text)
    matches: list[str] = []
    for family, signatures in DOCUMENT_FAMILY_SIGNATURES.items():
        if any(signature in normalized for signature in signatures):
            matches.append(family)
    return tuple(matches)


def split_citation_field(text: str) -> tuple[str, ...]:
    return unique_preserve_order(
        part.strip(" ;")
        for part in TOP_LEVEL_CITATION_SPLIT_RE.split(text.strip())
        if part.strip(" ;")
    )


def extract_point_values(text: str) -> tuple[str, ...]:
    normalized = normalize_for_matching(text)
    values: list[str] = []
    for match in LEGAL_POINT_GROUP_RE.finditer(normalized):
        raw_values = match.group("value")
        for point in re.split(r"\s*(?:,|va)\s*", raw_values):
            cleaned = point.strip()
            if cleaned:
                values.append(cleaned)
    return unique_preserve_order(values)


def compose_atomic_citation(
    *,
    article: str | None = None,
    clause: str | None = None,
    point: str | None = None,
    document_family: str | None = None,
) -> str:
    parts: list[str] = []
    if article:
        parts.append(f"Điều {article}")
    if clause:
        parts.append(f"khoản {clause}")
    if point:
        parts.append(f"điểm {point}")
    if document_family:
        parts.append(DOCUMENT_FAMILY_LABELS[document_family])
    return " ".join(parts).strip()


def expand_compound_citation(fragment: str, *, default_family: str | None = None) -> tuple[str, ...]:
    cleaned_fragment = fragment.strip()
    if not cleaned_fragment:
        return ()

    explicit_families = citation_document_families(cleaned_fragment)
    document_family = explicit_families[0] if len(explicit_families) == 1 else default_family
    normalized = normalize_signature_text(cleaned_fragment)
    article_values = unique_preserve_order(match.group("value") for match in LEGAL_ARTICLE_RE.finditer(normalized))
    clause_values = unique_preserve_order(match.group("value") for match in LEGAL_CLAUSE_RE.finditer(normalized))
    point_values = extract_point_values(cleaned_fragment)

    if len(article_values) > 1 and (clause_values or point_values):
        return (cleaned_fragment,)
    if len(clause_values) > 1 and point_values:
        return (cleaned_fragment,)

    if len(article_values) > 1:
        return unique_preserve_order(
            compose_atomic_citation(article=article, document_family=document_family)
            for article in article_values
        )

    if article_values:
        article = article_values[0]
        if point_values:
            clause = clause_values[0] if len(clause_values) == 1 else None
            if clause is None and clause_values:
                return (cleaned_fragment,)
            return unique_preserve_order(
                compose_atomic_citation(
                    article=article,
                    clause=clause,
                    point=point,
                    document_family=document_family,
                )
                for point in point_values
            )

        if len(clause_values) > 1:
            return unique_preserve_order(
                compose_atomic_citation(
                    article=article,
                    clause=clause,
                    document_family=document_family,
                )
                for clause in clause_values
            )

        return (
            compose_atomic_citation(
                article=article,
                clause=clause_values[0] if clause_values else None,
                document_family=document_family,
            ),
        )

    return (cleaned_fragment,)


def expand_expected_citations(values: Iterable[str | None]) -> tuple[str, ...]:
    expanded: list[str] = []
    for value in values:
        if not value:
            continue
        fragments = split_citation_field(value)
        families = citation_document_families(value)
        default_family = families[0] if len(families) == 1 else None
        for fragment in fragments:
            expanded.extend(expand_compound_citation(fragment, default_family=default_family))
    return unique_preserve_order(expanded)


def locate_header_row(rows: Sequence[Sequence[object]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        normalized = [normalize_for_matching(str(cell or "")) for cell in row]
        if normalized and normalized[0] == "id":
            return index, [str(cell or "").strip() for cell in row]
    raise ValueError("Could not locate benchmark header row.")


def parse_benchmark_rows(rows: Sequence[Sequence[object]]) -> list[BenchmarkCase]:
    header_index, headers = locate_header_row(rows)
    header_map = {header: position for position, header in enumerate(headers) if header}
    cases: list[BenchmarkCase] = []

    for row in rows[header_index + 1 :]:
        padded_row = list(row) + [None] * max(0, len(headers) - len(row))
        if is_placeholder_benchmark_row(padded_row):
            continue
        row_id = coerce_optional_text(padded_row[header_map["id"]])
        if row_id is None:
            continue

        def value_for(header: str) -> object:
            return padded_row[header_map[header]]

        def optional_value_for(header: str) -> object:
            if header not in header_map:
                return None
            return padded_row[header_map[header]]

        cases.append(
            BenchmarkCase(
                id=coerce_required_text(value_for("id"), "id"),
                category=coerce_required_text(value_for("category"), "category"),
                subtopic=coerce_required_text(value_for("subtopic"), "subtopic"),
                difficulty=coerce_required_text(value_for("difficulty"), "difficulty"),
                question_type=coerce_required_text(value_for("question_type"), "question_type"),
                question=coerce_required_text(value_for("question"), "question"),
                scenario=coerce_required_text(value_for("scenario"), "scenario"),
                gold_issue=coerce_required_text(value_for("gold_issue"), "gold_issue"),
                gold_citation_primary=coerce_optional_text(value_for("gold_citation_primary")),
                gold_citation_secondary=coerce_optional_text(value_for("gold_citation_secondary")),
                gold_answer_short=coerce_required_text(value_for("gold_answer_short"), "gold_answer_short"),
                gold_answer_full=coerce_required_text(value_for("gold_answer_full"), "gold_answer_full"),
                abstain_required=parse_yes_no_flag(value_for("abstain_required")),
                missing_information=coerce_optional_text(value_for("missing_information")),
                source_document=coerce_optional_text(value_for("source_document")),
                source_url=coerce_optional_text(value_for("source_url")),
                annotator=coerce_optional_text(value_for("annotator")),
                review_status=coerce_optional_text(value_for("review_status")),
                notes=coerce_optional_text(value_for("notes")),
                skill_tag=coerce_optional_text(optional_value_for("skill_tag")),
                gold_citations=expand_expected_citations(
                    (
                        coerce_optional_text(value_for("gold_citation_primary")),
                        coerce_optional_text(value_for("gold_citation_secondary")),
                    )
                ),
            )
        )

    return cases


def load_benchmark_workbook(
    workbook_path: Path,
    *,
    sheet_name: str = WORKBOOK_SHEET_NAME,
) -> list[BenchmarkCase]:
    load_workbook = require_openpyxl()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return parse_benchmark_rows(rows)


def load_benchmark_jsonl(input_path: Path) -> list[BenchmarkCase]:
    cases: list[BenchmarkCase] = []
    for line in input_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        payload = json.loads(line)
        cases.append(BenchmarkCase(**payload))
    return cases


def write_benchmark_jsonl(cases: Sequence[BenchmarkCase], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        for case in cases:
            handle.write(json.dumps(asdict(case), ensure_ascii=False) + "\n")


def result_columns(retrieval_hit_column: str = "retrieval_hit_at_5") -> tuple[str, ...]:
    return tuple(
        retrieval_hit_column if column == "retrieval_hit_at_5" else column
        for column in RESULTS_COLUMNS
    )


def write_results_csv(
    rows: Sequence[dict[str, object]],
    output_path: Path,
    *,
    fieldnames: Sequence[str] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames or RESULTS_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_results_jsonl(rows: Sequence[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def expected_citations(case: BenchmarkCase) -> tuple[str, ...]:
    return tuple(case.gold_citations)


def partition_citations_by_scope(
    citations: Sequence[str],
    *,
    allowed_document_families: Sequence[str] | None,
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    if not allowed_document_families:
        return tuple(citations), ()

    allowed = set(allowed_document_families)
    in_scope: list[str] = []
    out_of_scope: list[str] = []

    for citation in citations:
        family = citation_document_family(citation)
        if family is None or family in allowed:
            in_scope.append(citation)
        else:
            out_of_scope.append(citation)

    return tuple(in_scope), tuple(out_of_scope)


def expected_citations_in_scope(
    case: BenchmarkCase,
    *,
    allowed_document_families: Sequence[str] | None,
) -> tuple[str, ...]:
    in_scope, _ = partition_citations_by_scope(
        expected_citations(case),
        allowed_document_families=allowed_document_families,
    )
    return in_scope


def expected_citations_out_of_scope(
    case: BenchmarkCase,
    *,
    allowed_document_families: Sequence[str] | None,
) -> tuple[str, ...]:
    _, out_of_scope = partition_citations_by_scope(
        expected_citations(case),
        allowed_document_families=allowed_document_families,
    )
    return out_of_scope


def expected_citation_scope(
    case: BenchmarkCase,
    *,
    allowed_document_families: Sequence[str] | None,
) -> str:
    citations = expected_citations(case)
    if not citations:
        return "no_gold_citation"

    in_scope, out_of_scope = partition_citations_by_scope(
        citations,
        allowed_document_families=allowed_document_families,
    )
    if in_scope and out_of_scope:
        return "mixed_scope"
    if in_scope:
        return "in_scope"
    return "out_of_scope"


def document_families_from_chunk_paths(chunk_paths: Sequence[str | Path]) -> tuple[str, ...]:
    families: list[str] = []
    for chunk_path in chunk_paths:
        for family in citation_document_families(str(chunk_path)):
            families.append(family)
    return unique_preserve_order(families)


def citation_document_family(text: str) -> str | None:
    families = citation_document_families(text)
    if len(families) == 1:
        return families[0]
    return None


def resolve_citation_match_mode(mode: str | None = None) -> str:
    if mode is not None:
        return normalize_match_mode(mode)
    return normalize_match_mode(load_settings().eval_citation_match_mode)


def citation_matches_expected(expected: str, observed: str, *, mode: str | None = None) -> bool:
    expected_family = citation_document_family(expected)
    observed_family = citation_document_family(observed)
    if expected_family and observed_family and expected_family != observed_family:
        return False

    return citation_matches(
        retrieved_text=observed,
        expected_text=expected,
        mode=resolve_citation_match_mode(mode),
    )


def citation_document_matches_expected(expected: str, observed: str) -> bool:
    expected_family = citation_document_family(expected)
    observed_family = citation_document_family(observed)
    return bool(expected_family and observed_family and expected_family == observed_family)


def retrieval_hit_at_k(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    k: int = 5,
    allowed_document_families: Sequence[str] | None = None,
    citation_match_mode: str | None = None,
) -> bool | None:
    gold_citations = expected_citations_in_scope(
        case,
        allowed_document_families=allowed_document_families,
    )
    if not gold_citations:
        return None if allowed_document_families else False

    for expected in gold_citations:
        for observed in observed_citations[:k]:
            if citation_matches_expected(expected, observed, mode=citation_match_mode):
                return True
    return False


def score_citation_correctness(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    citation_match_mode: str | None = None,
) -> str:
    return score_citation_correctness_for_scope(
        case,
        observed_citations,
        citation_match_mode=citation_match_mode,
    )


def score_citation_correctness_for_scope(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    allowed_document_families: Sequence[str] | None = None,
    citation_match_mode: str | None = None,
) -> str:
    gold_citations = expected_citations_in_scope(
        case,
        allowed_document_families=allowed_document_families,
    )
    if not gold_citations:
        return "na" if allowed_document_families else "no"
    if not observed_citations:
        return "no"

    matched = [
        expected
        for expected in gold_citations
        if any(
            citation_matches_expected(expected, observed, mode=citation_match_mode)
            for observed in observed_citations
        )
    ]
    if len(matched) == len(gold_citations):
        return "exact"
    if matched:
        return "partial"
    return "no"


def score_citation_article_correctness_for_scope(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    allowed_document_families: Sequence[str] | None = None,
    citation_match_mode: str | None = None,
) -> str:
    return score_citation_correctness_for_scope(
        case,
        observed_citations,
        allowed_document_families=allowed_document_families,
        citation_match_mode=citation_match_mode,
    )


def score_citation_document_correctness_for_scope(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    allowed_document_families: Sequence[str] | None = None,
) -> str:
    gold_citations = expected_citations_in_scope(
        case,
        allowed_document_families=allowed_document_families,
    )
    expected_families = unique_preserve_order(
        family
        for citation in gold_citations
        for family in (citation_document_family(citation),)
        if family
    )
    if not expected_families:
        return "na"
    if not observed_citations:
        return "no"

    observed_families = {
        family
        for citation in observed_citations
        for family in (citation_document_family(citation),)
        if family
    }
    matched = [family for family in expected_families if family in observed_families]
    if len(matched) == len(expected_families):
        return "exact"
    if matched:
        return "partial"
    return "no"


def first_relevant_rank(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    k: int = 5,
    allowed_document_families: Sequence[str] | None = None,
    citation_match_mode: str | None = None,
) -> int | None:
    gold_citations = expected_citations_in_scope(
        case,
        allowed_document_families=allowed_document_families,
    )
    if not gold_citations:
        return None if allowed_document_families else 0

    for rank, observed in enumerate(observed_citations[:k], start=1):
        if any(
            citation_matches_expected(expected, observed, mode=citation_match_mode)
            for expected in gold_citations
        ):
            return rank
    return 0


def reciprocal_rank(
    case: BenchmarkCase,
    observed_citations: Sequence[str],
    *,
    k: int = 5,
    allowed_document_families: Sequence[str] | None = None,
    citation_match_mode: str | None = None,
) -> float | None:
    rank = first_relevant_rank(
        case,
        observed_citations,
        k=k,
        allowed_document_families=allowed_document_families,
        citation_match_mode=citation_match_mode,
    )
    if rank is None:
        return None
    if rank <= 0:
        return 0.0
    return 1.0 / rank


def mean_reciprocal_rank(reciprocal_ranks: Sequence[float | None]) -> float | None:
    scored = [rank for rank in reciprocal_ranks if rank is not None]
    if not scored:
        return None
    return sum(scored) / len(scored)


def case_requires_missing_information_handling(case: BenchmarkCase) -> bool:
    return case.abstain_required or bool(case.missing_information)


def yes_partial_no_points(value: str, *, full_points: int, partial_points: int) -> int:
    normalized = value.strip().lower()
    if normalized == "yes":
        return full_points
    if normalized == "partial":
        return partial_points
    return 0


def exact_partial_no_points(value: str, *, full_points: int, partial_points: int) -> int:
    normalized = value.strip().lower()
    if normalized == "exact":
        return full_points
    if normalized == "partial":
        return partial_points
    return 0


def legal_reasoning_points(score: int) -> int:
    if score >= 4:
        return 2
    if score >= 2:
        return 1
    return 0


def groundedness_points(score: int) -> int:
    return 1 if score >= 3 else 0


def clarity_format_points(clarity_score: int, format_score: int) -> int:
    return 1 if (clarity_score + format_score) / 2 >= 4 else 0


def compute_final_score_10(
    *,
    case: BenchmarkCase,
    answer_correct: str,
    legal_issue_classification_correct: str,
    legal_reasoning_score_1_5: int,
    missing_information_score_0_2: int,
    citation_article_correct: str,
    citation_supports_answer: str,
    groundedness_score_1_5: int,
    clarity_score_1_5: int,
    format_score_1_5: int,
) -> int:
    score = 0
    score += yes_partial_no_points(
        legal_issue_classification_correct,
        full_points=2,
        partial_points=1,
    )
    score += yes_partial_no_points(answer_correct, full_points=2, partial_points=1)
    if case_requires_missing_information_handling(case):
        score += max(0, min(2, missing_information_score_0_2))
    else:
        score += legal_reasoning_points(legal_reasoning_score_1_5)
    citation_article_points = exact_partial_no_points(
        citation_article_correct,
        full_points=2,
        partial_points=1,
    )
    citation_support_points = yes_partial_no_points(
        citation_supports_answer,
        full_points=2,
        partial_points=1,
    )
    score += min(citation_article_points, citation_support_points)
    score += groundedness_points(groundedness_score_1_5)
    score += clarity_format_points(clarity_score_1_5, format_score_1_5)
    return max(0, min(10, score))


def build_judge_messages(
    case: BenchmarkCase,
    *,
    generated_answer: str,
    generated_legal_basis: Sequence[str],
    insufficient_context: str,
    expected_citations_scoped: Sequence[str],
    retrieved_citations: Sequence[str],
    case_scope: str,
) -> list[dict[str, str]]:
    user_prompt = "\n\n".join(
        [
            f"QUESTION:\n{case.question.strip()}",
            f"GOLD_ISSUE:\n{case.gold_issue.strip()}",
            f"GOLD_ANSWER:\n{case.gold_answer_full.strip()}",
            f"GOLD_ANSWER_SHORT:\n{case.gold_answer_short.strip()}",
            f"ABSTAIN_REQUIRED:\n{case.abstain_required}",
            f"GOLD_MISSING_INFORMATION:\n{(case.missing_information or '').strip()}",
            f"CASE_SCOPE:\n{case_scope}",
            "EXPECTED_CITATIONS_IN_SCOPE:",
            "\n".join(f"- {citation}" for citation in expected_citations_scoped) or "-",
            "Judge chi duoc dua tren evidence duoc cap trong prompt nay.",
            "RETRIEVED_CITATIONS:",
            "\n".join(f"- {citation}" for citation in retrieved_citations) or "-",
            f"GENERATED_ANSWER:\n{generated_answer.strip()}",
            "GENERATED_LEGAL_BASIS:",
            "\n".join(f"- {citation}" for citation in generated_legal_basis) or "-",
            f"INSUFFICIENT_CONTEXT:\n{insufficient_context}",
        ]
    )
    return [
        {
            "role": "system",
            "content": JUDGE_SYSTEM_PROMPT,
        },
        {"role": "user", "content": user_prompt},
    ]


def extract_json_candidate(raw_content: str) -> str:
    cleaned_content = raw_content.strip()

    fenced_match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", cleaned_content, re.IGNORECASE)
    if fenced_match:
        cleaned_content = fenced_match.group(1).strip()

    if cleaned_content.startswith("{") and cleaned_content.endswith("}"):
        return cleaned_content

    container_match = re.search(r"(\{.*\})", cleaned_content, re.DOTALL)
    if container_match:
        return container_match.group(1).strip()

    return cleaned_content


def parse_hallucination_types(value: object) -> tuple[str, ...] | None:
    if value is None:
        return ()

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"", "none", "no", "[]"}:
            return ()
        raw_values = re.split(r"\s*(?:,|;|\|)\s*", normalized)
    elif isinstance(value, list):
        raw_values = [str(item).strip().lower() for item in value]
    else:
        return None

    allowed = {"legal_basis", "rule", "fact"}
    parsed: list[str] = []
    for raw_value in raw_values:
        if not raw_value:
            continue
        if raw_value not in allowed:
            return None
        parsed.append(raw_value)
    return unique_preserve_order(parsed)


def parse_judge_payload(raw_content: str) -> JudgeScore | None:
    cleaned_content = extract_json_candidate(raw_content)
    try:
        payload = json.loads(cleaned_content)
    except json.JSONDecodeError:
        return None

    if not isinstance(payload, dict):
        return None

    answer_correct = str(payload.get("answer_correct") or "").strip().lower()
    if answer_correct not in {"yes", "partial", "no"}:
        return None

    legal_issue_classification_correct = str(
        payload.get("legal_issue_classification_correct") or ""
    ).strip().lower()
    if legal_issue_classification_correct not in {"yes", "partial", "no"}:
        return None

    citation_supports_answer = str(payload.get("citation_supports_answer") or "").strip().lower()
    if citation_supports_answer not in {"yes", "partial", "no"}:
        return None

    hallucination_types = parse_hallucination_types(payload.get("hallucination_types"))
    if hallucination_types is None:
        return None

    try:
        legal_reasoning_score = int(payload["legal_reasoning_score_1_5"])
        missing_information_score = int(payload["missing_information_score_0_2"])
        groundedness_score = int(payload["groundedness_score_1_5"])
        clarity_score = int(payload["clarity_score_1_5"])
        format_score = int(payload["format_score_1_5"])
    except (KeyError, TypeError, ValueError):
        return None

    if not (
        1 <= legal_reasoning_score <= 5
        and 0 <= missing_information_score <= 2
        and 1 <= groundedness_score <= 5
        and 1 <= clarity_score <= 5
        and 1 <= format_score <= 5
    ):
        return None

    return JudgeScore(
        answer_correct=answer_correct,
        legal_issue_classification_correct=legal_issue_classification_correct,
        legal_reasoning_score_1_5=legal_reasoning_score,
        missing_information_score_0_2=missing_information_score,
        citation_supports_answer=citation_supports_answer,
        groundedness_score_1_5=groundedness_score,
        clarity_score_1_5=clarity_score,
        format_score_1_5=format_score,
        hallucination_types=hallucination_types,
        comments=str(payload.get("comments") or "").strip(),
        raw_content=raw_content,
    )


def summarize_benchmark_cases(cases: Sequence[BenchmarkCase]) -> dict[str, object]:
    difficulties: dict[str, int] = {}
    categories: dict[str, int] = {}
    for case in cases:
        difficulties[case.difficulty] = difficulties.get(case.difficulty, 0) + 1
        categories[case.category] = categories.get(case.category, 0) + 1
    return {
        "case_count": len(cases),
        "difficulty_distribution": difficulties,
        "category_distribution": categories,
    }


__all__ = [
    "BENCHMARK_JSONL_NAME",
    "BenchmarkCase",
    "RESULTS_COLUMNS",
    "WORKBOOK_RESULTS_SHEET_NAME",
    "WORKBOOK_SHEET_NAME",
    "CitationRef",
    "case_requires_missing_information_handling",
    "citation_contains",
    "citation_matches_expected",
    "citation_document_matches_expected",
    "document_families_from_chunk_paths",
    "JUDGE_JSON_SCHEMA",
    "JudgeScore",
    "build_judge_messages",
    "compute_final_score_10",
    "expected_citations",
    "expected_citations_in_scope",
    "expected_citations_out_of_scope",
    "expected_citation_scope",
    "first_relevant_rank",
    "load_benchmark_jsonl",
    "load_benchmark_workbook",
    "mean_reciprocal_rank",
    "parse_benchmark_rows",
    "parse_yes_no_flag",
    "parse_hallucination_types",
    "parse_judge_payload",
    "partition_citations_by_scope",
    "parse_citation",
    "reciprocal_rank",
    "require_openpyxl",
    "resolve_citation_match_mode",
    "result_columns",
    "retrieval_hit_at_k",
    "score_citation_article_correctness_for_scope",
    "score_citation_correctness",
    "score_citation_correctness_for_scope",
    "score_citation_document_correctness_for_scope",
    "summarize_benchmark_cases",
    "normalize_vietnamese_citation",
    "write_benchmark_jsonl",
    "write_results_csv",
    "write_results_jsonl",
]
